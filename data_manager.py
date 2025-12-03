import pandas as pd
import openpyxl
import os
import time
import google_sheets_manager


CREDENTIALS_FILE = "credentials.json"
SPREADSHEET_NAME = "https://docs.google.com/spreadsheets/d/1477Q9GMKGSzMlWp6ao7YSt0CYXOsVPx87K8rbDdCfVM/edit?gid=0#gid=0" # Default name
FILE_PATH = "recruitment_data.xlsx"  # Default Excel file path (fallback if Google Sheets not available)

# Mapping from Excel Header to App Internal Name - EXPANDED
COLUMN_MAPPING = {
    "RECRUITMENT PHASE\n(Pipeline)": "STATUS",
    "JOB APPLIED FOR": "POSITION",
    "APPLIED DATE": "APPLICATION DATE",
    "CANDIDATE NAME": "CANDIDATE NAME",
    "RECRUITER": "RECRUITER",
    "SOURCE": "SOURCE",
    "EMAIL": "EMAIL", 
    "PHONE": "PHONE", 
    "COMMENTS": "NOTES",
    "JOB ID": "JOB ID",
    "DEPARTMENT": "DEPARTMENT",
    "FINAL DECISION": "FINAL DECISION",
    "HR VIEW": "HR VIEW",
    "HIRING MANAGER VIEW": "HIRING MANAGER VIEW",
    "DECISION MAKER VIEW": "DECISION MAKER VIEW",
    "RECIEVED APPLICATION COMMENTS": "RECEIVED APPLICATION COMMENTS",
    "RECEIVED APPLICATION COMMENTS": "RECEIVED APPLICATION COMMENTS"
}

# Job Openings column mapping
JOB_MAPPING = {
    "JOB ID": "JOB ID",
    "DEPARTMENT": "DEPARTMENT",
    "JOB TITLE": "JOB TITLE",
    "OPENING DATE": "OPENING DATE",
    "RECRUITER": "RECRUITER",
    "STATUS": "STATUS",
    "NEW HIRE START DATE": "NEW HIRE START DATE",
    "HIRING COST": "HIRING COST"
}

# Reverse mapping for saving
REVERSE_MAPPING = {v: k for k, v in COLUMN_MAPPING.items()}
REVERSE_JOB_MAPPING = {v: k for k, v in JOB_MAPPING.items()}

class DataManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.candidates_df = pd.DataFrame()
        self.jobs_df = pd.DataFrame()
        self.preferences = {}
        self.last_load_time = 0
        
        # Google Sheets integration
        # Check if running on Streamlit Cloud with secrets
        try:
            import streamlit as st
            if hasattr(st, 'secrets') and 'SPREADSHEET_URL' in st.secrets:
                spreadsheet_url = st.secrets['SPREADSHEET_URL']
                self.use_google_sheets = True
                self.gs_manager = google_sheets_manager.GoogleSheetManager(CREDENTIALS_FILE, spreadsheet_url)
            elif os.path.exists(CREDENTIALS_FILE):
                self.use_google_sheets = True
                self.gs_manager = google_sheets_manager.GoogleSheetManager(CREDENTIALS_FILE, SPREADSHEET_NAME)
            else:
                self.use_google_sheets = False
                self.gs_manager = None
        except:
            # Fallback if streamlit import fails or no secrets
            self.use_google_sheets = os.path.exists(CREDENTIALS_FILE)
            self.gs_manager = None
            if self.use_google_sheets:
                self.gs_manager = google_sheets_manager.GoogleSheetManager(CREDENTIALS_FILE, SPREADSHEET_NAME)
        
    def _get_header_row(self, df, target_column="CANDIDATE NAME"):
        """Finds the row index containing the target column."""
        # Optimize: Only check first 20 rows
        for i, row in df.head(20).iterrows():
            row_values = [str(val).strip().upper() for val in row.values]
            if target_column in row_values:
                return i
        return None

    def load_data(self):
        """Loads all necessary data from the Excel file or Google Sheets."""
        
        # Try Google Sheets first if configured
        if self.use_google_sheets and self.gs_manager:
            success, msg = self.gs_manager.load_data()
            if success:
                self.candidates_df = self.gs_manager.candidates_df
                self.jobs_df = self.gs_manager.jobs_df
                self.preferences = self.gs_manager.preferences
                self.last_load_time = time.time()
                return True, f"Dati caricati da Google Sheets"
            else:
                # If Google Sheets fails and no local file exists, return detailed error
                if not os.path.exists(self.file_path):
                    return False, f"❌ Google Sheets Error: {msg}\n\n💡 Possibili soluzioni:\n1. Verifica che il foglio sia condiviso con: dylan4luigi@papa-479815.iam.gserviceaccount.com\n2. Controlla che i Secrets siano configurati correttamente su Streamlit Cloud\n3. Verifica che l'URL del foglio sia corretto"
                # Otherwise fallback to local file
                print(f"⚠️ Google Sheets fallito ({msg}), provo con file locale...")

        # Use local Excel file
        if not os.path.exists(self.file_path):
            return False, f"❌ File Excel '{self.file_path}' non trovato.\n\n💡 Per usare Google Sheets:\n1. Configura i Secrets su Streamlit Cloud\n2. Condividi il foglio con: dylan4luigi@papa-479815.iam.gserviceaccount.com"

        try:
            # Use ExcelFile to open once and read multiple sheets
            xls = pd.ExcelFile(self.file_path, engine='openpyxl')
            
            # 1. Load Candidates
            if "Candidates" in xls.sheet_names:
                df_raw = pd.read_excel(xls, sheet_name="Candidates", header=None, nrows=20)
                header_row = self._get_header_row(df_raw, "CANDIDATE NAME")
                
                if header_row is not None:
                    self.candidates_df = pd.read_excel(xls, sheet_name="Candidates", header=header_row)
                    self.candidates_df = self.candidates_df.dropna(subset=["CANDIDATE NAME"])
                    self.candidates_df = self.candidates_df.loc[:, ~self.candidates_df.columns.str.contains('^Unnamed')]
                    self.candidates_df = self.candidates_df.rename(columns=COLUMN_MAPPING)
                    
                    # Ensure critical columns exist
                    for col in ["STATUS", "POSITION", "APPLICATION DATE", "RECRUITER", "SOURCE", 
                               "JOB ID", "DEPARTMENT", "FINAL DECISION", "HR VIEW", 
                               "HIRING MANAGER VIEW", "DECISION MAKER VIEW", "RECEIVED APPLICATION COMMENTS"]:
                        if col not in self.candidates_df.columns:
                            self.candidates_df[col] = None
            
            # 2. Load Job Openings
            if "JobOpenings" in xls.sheet_names:
                df_jobs_raw = pd.read_excel(xls, sheet_name="JobOpenings", header=None, nrows=20)
                job_header = self._get_header_row(df_jobs_raw, "JOB ID")
                
                if job_header is not None:
                    self.jobs_df = pd.read_excel(xls, sheet_name="JobOpenings", header=job_header)
                    # Remove rows where JOB ID is NaN
                    self.jobs_df = self.jobs_df.dropna(subset=["JOB ID"])
                    self.jobs_df = self.jobs_df.loc[:, ~self.jobs_df.columns.str.contains('^Unnamed')]
                    self.jobs_df = self.jobs_df.rename(columns=JOB_MAPPING)
                    
                    # Ensure critical columns
                    for col in ["JOB ID", "DEPARTMENT", "JOB TITLE", "OPENING DATE", 
                               "RECRUITER", "STATUS", "NEW HIRE START DATE", "HIRING COST"]:
                        if col not in self.jobs_df.columns:
                            self.jobs_df[col] = None
            
            # 3. Load Preferences & extract lists
            self.preferences = {
                "Status": [],
                "Recruiters": [],
                "Sources": [],
                "Positions": [],
                "Departments": [],
                "Decision Comments": [],
                "Job Statuses": ["Vacant", "Filled", "Suspended", "Cancelled"]
            }
            
            if "Preferences" in xls.sheet_names:
                df_prefs = pd.read_excel(xls, sheet_name="Preferences", header=None)
                # Helper to get values from a column starting at row 7
                def get_col_values(col_idx, start_row=7):
                    if col_idx < df_prefs.shape[1]:
                        values = df_prefs.iloc[start_row:, col_idx].dropna().astype(str).tolist()
                        return [v for v in values if v.lower() != "nan" and v.strip() != ""]
                    return []

                self.preferences["Recruiters"] = get_col_values(2)
                self.preferences["Status"] = get_col_values(11)
                
                # Try to get decision comments (usually around column 11-12, rows after status)
                # This is an approximation based on the template structure
                decision_comments_raw = get_col_values(11, start_row=18)
                if decision_comments_raw:
                    self.preferences["Decision Comments"] = decision_comments_raw

            # Extract departments from jobs
            if not self.jobs_df.empty and "DEPARTMENT" in self.jobs_df.columns:
                self.preferences["Departments"] = self.jobs_df["DEPARTMENT"].dropna().unique().tolist()
            
            # Extract positions from jobs
            if not self.jobs_df.empty and "JOB TITLE" in self.jobs_df.columns:
                self.preferences["Positions"] = self.jobs_df["JOB TITLE"].dropna().unique().tolist()
            
            # Fallback for Sources from candidates
            if "SOURCE" in self.candidates_df.columns:
                current_sources = self.candidates_df["SOURCE"].dropna().unique().tolist()
                self.preferences["Sources"] = list(set(self.preferences.get("Sources", []) + current_sources))
                
            self.last_load_time = time.time()
            return True, "Dati caricati con successo (Locale)."

        except PermissionError:
            return False, "Il file Excel è aperto in un altro programma. Chiudilo e riprova."
        except Exception as e:
            return False, f"Errore nel caricamento dati: {e}"

    def save_candidate(self, candidate_dict):
        """Appends a new candidate to the Excel file or Google Sheets."""
        if self.use_google_sheets:
            return self.gs_manager.save_candidate(candidate_dict)
            
        try:
            # Check file lock first
            try:
                with open(self.file_path, "a"):
                    pass
            except PermissionError:
                return False, "Il file Excel è aperto in un altro programma. Chiudilo per salvare."

            df_raw = pd.read_excel(self.file_path, sheet_name="Candidates", header=None, nrows=20)
            header_row = self._get_header_row(df_raw, "CANDIDATE NAME")
            
            if header_row is None:
                return False, "Impossibile trovare l'intestazione nel foglio Candidates."

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Candidates"]
            
            # Get columns from the dataframe read with correct header
            df_header = pd.read_excel(self.file_path, sheet_name="Candidates", header=header_row, nrows=0)
            df_cols = df_header.columns.tolist()
            
            new_row = []
            for col in df_cols:
                # Map back to internal name if possible
                mapped_key = COLUMN_MAPPING.get(col, None) 
                
                val = None
                if mapped_key and mapped_key in candidate_dict:
                    val = candidate_dict[mapped_key]
                elif col in candidate_dict: # Direct match
                    val = candidate_dict[col]
                
                new_row.append(val)
                
            ws.append(new_row)
            wb.save(self.file_path)
            
            # Reload data to keep cache in sync
            self.load_data()
            
            return True, "Candidato salvato con successo!"
        except PermissionError:
             return False, "Il file Excel è aperto in un altro programma. Chiudilo per salvare."
        except Exception as e:
            return False, f"Errore nel salvataggio: {e}"

    def save_job_opening(self, job_dict):
        """Appends a new job opening to the Excel file or Google Sheets."""
        if self.use_google_sheets:
            return self.gs_manager.save_job_opening(job_dict)
            
        try:
            # Check file lock first
            try:
                with open(self.file_path, "a"):
                    pass
            except PermissionError:
                return False, "Il file Excel è aperto in un altro programma. Chiudilo per salvare."

            df_raw = pd.read_excel(self.file_path, sheet_name="JobOpenings", header=None, nrows=20)
            header_row = self._get_header_row(df_raw, "JOB ID")
            
            if header_row is None:
                return False, "Impossibile trovare l'intestazione nel foglio JobOpenings."

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["JobOpenings"]
            
            # Get columns
            df_header = pd.read_excel(self.file_path, sheet_name="JobOpenings", header=header_row, nrows=0)
            df_cols = df_header.columns.tolist()
            
            new_row = []
            for col in df_cols:
                mapped_key = JOB_MAPPING.get(col, None)
                
                val = None
                if mapped_key and mapped_key in job_dict:
                    val = job_dict[mapped_key]
                elif col in job_dict:
                    val = job_dict[col]
                
                new_row.append(val)
                
            ws.append(new_row)
            wb.save(self.file_path)
            
            # Reload data
            self.load_data()
            
            return True, "Posizione salvata con successo!"
        except PermissionError:
             return False, "Il file Excel è aperto in un altro programma. Chiudilo per salvare."
        except Exception as e:
            return False, f"Errore nel salvataggio: {e}"

    def update_candidate(self, candidate_name, updates):
        """Updates an existing candidate.
        
        Args:
            candidate_name: Nome del candidato da aggiornare
            updates: Dizionario con i campi da aggiornare {field_name: new_value}
            
        Returns:
            Tuple (success: bool, message: str)
        """
        if self.use_google_sheets:
            return self.gs_manager.update_candidate(candidate_name, updates)
            
        try:
            # Check file lock first
            try:
                with open(self.file_path, "a"):
                    pass
            except PermissionError:
                return False, "Il file Excel è aperto in un altro programma. Chiudilo per aggiornare."

            # Excel implementation
            df_raw = pd.read_excel(self.file_path, sheet_name="Candidates", header=None, nrows=20)
            header_row = self._get_header_row(df_raw, "CANDIDATE NAME")
            
            if header_row is None:
                return False, "Impossibile trovare l'intestazione nel foglio Candidates."

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Candidates"]
            
            # Get headers from the Excel file
            df_header = pd.read_excel(self.file_path, sheet_name="Candidates", header=header_row, nrows=0)
            headers = df_header.columns.tolist()
            
            # Find candidate name column index
            name_col_idx = None
            for idx, col in enumerate(headers):
                mapped_name = COLUMN_MAPPING.get(col, col)
                if mapped_name == "CANDIDATE NAME" or col == "CANDIDATE NAME":
                    name_col_idx = idx
                    break
            
            if name_col_idx is None:
                return False, "Colonna CANDIDATE NAME non trovata."
            
            # Find the row with this candidate (1-based for openpyxl)
            found_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+2), start=header_row+2):
                cell_value = row[name_col_idx].value
                if cell_value and str(cell_value).strip() == str(candidate_name).strip():
                    found_row = row_idx
                    break
            
            if found_row is None:
                return False, f"Candidato '{candidate_name}' non trovato."
            
            # Update cells based on the updates dictionary
            for field, value in updates.items():
                # Find the Excel column name for this field
                # First check if field is already an Excel column name
                excel_col = None
                if field in headers:
                    excel_col = field
                else:
                    # Try reverse mapping (internal name -> Excel name)
                    excel_col = REVERSE_MAPPING.get(field, None)
                
                if excel_col and excel_col in headers:
                    col_idx = headers.index(excel_col)
                    # openpyxl uses 1-based indexing
                    ws.cell(row=found_row, column=col_idx+1, value=value)
            
            wb.save(self.file_path)
            self.load_data()
            
            return True, "Candidato aggiornato con successo!"
        except PermissionError:
            return False, "Il file Excel è aperto in un altro programma. Chiudilo per aggiornare."
        except Exception as e:
            return False, f"Errore nell'aggiornamento: {e}"

    def update_job_opening(self, job_id, updates):
        """Updates an existing job opening.
        
        Args:
            job_id: ID della posizione da aggiornare
            updates: Dizionario con i campi da aggiornare {field_name: new_value}
            
        Returns:
            Tuple (success: bool, message: str)
        """
        if self.use_google_sheets:
            return self.gs_manager.update_job_opening(job_id, updates)
            
        try:
            # Check file lock first
            try:
                with open(self.file_path, "a"):
                    pass
            except PermissionError:
                return False, "Il file Excel è aperto in un altro programma. Chiudilo per aggiornare."

            df_raw = pd.read_excel(self.file_path, sheet_name="JobOpenings", header=None, nrows=20)
            header_row = self._get_header_row(df_raw, "JOB ID")
            
            if header_row is None:
                return False, "Impossibile trovare l'intestazione nel foglio JobOpenings."

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["JobOpenings"]
            
            # Get headers
            df_header = pd.read_excel(self.file_path, sheet_name="JobOpenings", header=header_row, nrows=0)
            headers = df_header.columns.tolist()
            
            # Find JOB ID column index
            job_id_col_idx = None
            for idx, col in enumerate(headers):
                mapped_name = JOB_MAPPING.get(col, col)
                if mapped_name == "JOB ID" or col == "JOB ID":
                    job_id_col_idx = idx
                    break
            
            if job_id_col_idx is None:
                return False, "Colonna JOB ID non trovata."
            
            # Find the row with this job ID
            found_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+2), start=header_row+2):
                cell_value = row[job_id_col_idx].value
                # Convert both to string for comparison
                if cell_value is not None and str(cell_value).strip() == str(job_id).strip():
                    found_row = row_idx
                    break
            
            if found_row is None:
                return False, f"Posizione con ID '{job_id}' non trovata."
            
            # Update cells based on the updates dictionary
            for field, value in updates.items():
                # Find the Excel column name for this field
                excel_col = None
                if field in headers:
                    excel_col = field
                else:
                    # Try reverse mapping (internal name -> Excel name)
                    excel_col = REVERSE_JOB_MAPPING.get(field, None)
                
                if excel_col and excel_col in headers:
                    col_idx = headers.index(excel_col)
                    # openpyxl uses 1-based indexing
                    ws.cell(row=found_row, column=col_idx+1, value=value)
            
            wb.save(self.file_path)
            self.load_data()
            
            return True, "Posizione aggiornata con successo!"
        except PermissionError:
            return False, "Il file Excel è aperto in un altro programma. Chiudilo per aggiornare."
        except Exception as e:
            return False, f"Errore nell'aggiornamento: {e}"

# Global instance
manager = DataManager(FILE_PATH)

def load_data():
    return manager.load_data()

def get_candidates():
    return manager.candidates_df

def get_jobs():
    return manager.jobs_df

def get_preferences():
    return manager.preferences

def save_candidate(data):
    return manager.save_candidate(data)

def save_job_opening(data):
    return manager.save_job_opening(data)

def update_candidate(name, updates):
    return manager.update_candidate(name, updates)

def update_job_opening(job_id, updates):
    return manager.update_job_opening(job_id, updates)
