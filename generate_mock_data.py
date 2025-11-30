"""
Mock Data Generator for Recruitment Tracker
Creates realistic sample data for testing
"""

import pandas as pd
from datetime import datetime, timedelta
import random
import openpyxl

# Sample data
RECRUITERS = ["Austin Millfrey", "Peter Caron", "Wesley Crafter", "Sarah Johnson", "Mike Thompson"]
DEPARTMENTS = ["Sales", "Marketing", "Human Resources", "IT", "Finance", "Operations", "Legal Affairs", "Public Relations"]
JOB_TITLES = {
    "Sales": ["Sales Specialist", "Account Manager", "Sales Director"],
    "Marketing": ["Marketing Manager", "Content Creator", "SEO Specialist"],
    "Human Resources": ["HR Specialist", "HR Manager", "Recruiter"],
    "IT": ["Software Developer", "IT Support", "DevOps Engineer"],
    "Finance": ["Financial Analyst", "Accountant", "CFO"],
    "Operations": ["Operations Manager", "Logistics Coordinator"],
    "Legal Affairs": ["Attorney", "Legal Counsel"],
    "Public Relations": ["PR Manager", "Social Media Manager"]
}
SOURCES = ["LinkedIn", "Job Portals", "Own Website", "Recruitment Agency", "Facebook", "Twitter", "Instagram", "Referral", "Newspaper"]
STATUSES = ["Received Application", "Sent to Manager", "Interviews", "Tests", "Job Offer", "Hired"]
FINAL_DECISIONS = ["Hired", "Not Hired", "Candidate in Process", "Candidate Refusal", ""]
JOB_STATUSES = ["Vacant", "Filled", "Suspended", "Cancelled"]

FIRST_NAMES = ["John", "Emma", "Michael", "Sophia", "William", "Olivia", "James", "Ava", "Robert", "Isabella",
               "David", "Mia", "Richard", "Charlotte", "Joseph", "Amelia", "Thomas", "Harper", "Charles", "Evelyn"]
LAST_NAMES = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez",
              "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"]

def generate_job_openings(num_jobs=15):
    """Generate mock job openings"""
    jobs = []
    
    for i in range(1, num_jobs + 1):
        dept = random.choice(DEPARTMENTS)
        job_title = random.choice(JOB_TITLES.get(dept, ["Specialist"]))
        
        opening_date = datetime.now() - timedelta(days=random.randint(30, 365))
        status = random.choice(JOB_STATUSES)
        
        # If filled, add start date
        start_date = None
        if status == "Filled":
            start_date = opening_date + timedelta(days=random.randint(30, 90))
        
        hiring_cost = random.randint(100, 5000)
        
        jobs.append({
            "JOB ID": i,
            "DEPARTMENT": dept,
            "JOB TITLE": job_title,
            "OPENING DATE": opening_date,
            "RECRUITER": random.choice(RECRUITERS),
            "STATUS": status,
            "NEW HIRE START DATE": start_date,
            "HIRING COST": hiring_cost
        })
    
    return pd.DataFrame(jobs)

def generate_candidates(num_candidates=30, jobs_df=None):
    """Generate mock candidates"""
    candidates = []
    
    for i in range(num_candidates):
        first_name = random.choice(FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)
        name = f"{first_name} {last_name}"
        
        email = f"{first_name.lower()}.{last_name.lower()}@email.com"
        phone = f"+1-{random.randint(100,999)}-{random.randint(100,999)}-{random.randint(1000,9999)}"
        
        application_date = datetime.now() - timedelta(days=random.randint(1, 180))
        
        # Select a job
        if jobs_df is not None and not jobs_df.empty:
            job = jobs_df.sample(1).iloc[0]
            job_id = job["JOB ID"]
            position = job["JOB TITLE"]
            department = job["DEPARTMENT"]
        else:
            job_id = random.randint(1, 10)
            department = random.choice(DEPARTMENTS)
            position = random.choice(JOB_TITLES.get(department, ["Specialist"]))
        
        recruiter = random.choice(RECRUITERS)
        source = random.choice(SOURCES)
        status = random.choice(STATUSES)
        
        # Final decision based on status
        if status == "Hired":
            final_decision = "Hired"
        elif status in ["Received Application", "Sent to Manager"]:
            final_decision = "Candidate in Process"
        else:
            final_decision = random.choice(FINAL_DECISIONS)
        
        # Generate stakeholder views
        hr_view = random.choice([
            "Strong technical skills",
            "Good cultural fit",
            "Excellent communication",
            "Needs more experience",
            "Perfect match for the role",
            ""
        ])
        
        manager_view = random.choice([
            "Impressed with portfolio",
            "Good problem-solving skills",
            "Team player",
            "Lacks specific expertise",
            "Highly recommended",
            ""
        ])
        
        decision_maker_view = random.choice([
            "Approve for hire",
            "Request second interview",
            "Salary expectations too high",
            "Strong candidate",
            "Not a fit",
            ""
        ])
        
        received_comments = random.choice([
            "Resume looks promising",
            "Referred by employee",
            "Applied through LinkedIn",
            "Direct application",
            ""
        ])
        
        notes = random.choice([
            "Follow up in 2 weeks",
            "Schedule technical interview",
            "Waiting for references",
            "Offer sent",
            "Declined offer",
            ""
        ])
        
        candidates.append({
            "JOB ID": job_id,
            "DEPARTMENT": department,
            "JOB APPLIED FOR": position,
            "RECRUITER": recruiter,
            "CANDIDATE NAME": name,
            "SOURCE": source,
            "APPLIED DATE": application_date,
            "RECRUITMENT PHASE\n(Pipeline)": status,
            "FINAL DECISION": final_decision,
            "HR VIEW": hr_view,
            "HIRING MANAGER VIEW": manager_view,
            "DECISION MAKER VIEW": decision_maker_view,
            "COMMENTS": notes,
            "RECIEVED APPLICATION COMMENTS": received_comments,
            "EMAIL": email,
            "PHONE": phone
        })
    
    return pd.DataFrame(candidates)

def add_mock_data_to_excel(file_path, num_jobs=15, num_candidates=30):
    """Add mock data to the Excel file"""
    print(f"Adding mock data to {file_path}...")
    
    # Generate data
    jobs_df = generate_job_openings(num_jobs)
    candidates_df = generate_candidates(num_candidates, jobs_df)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Add to JobOpenings sheet
    if "JobOpenings" in wb.sheetnames:
        ws_jobs = wb["JobOpenings"]
        
        # Find header row (assuming row 5 based on template)
        header_row = 6  # Row 6 in Excel (index 5 in Python)
        
        # Append jobs
        for _, job in jobs_df.iterrows():
            ws_jobs.append([
                None,  # Empty first column
                job["JOB ID"],
                job["DEPARTMENT"],
                job["JOB TITLE"],
                job["OPENING DATE"],
                job["RECRUITER"],
                job["STATUS"],
                job["NEW HIRE START DATE"],
                job["HIRING COST"]
            ])
        
        print(f"Added {len(jobs_df)} job openings")
    
    # Add to Candidates sheet
    if "Candidates" in wb.sheetnames:
        ws_candidates = wb["Candidates"]
        
        # Append candidates
        for _, candidate in candidates_df.iterrows():
            ws_candidates.append([
                None,  # Empty first column
                candidate["JOB ID"],
                candidate["DEPARTMENT"],
                candidate["JOB APPLIED FOR"],
                candidate["RECRUITER"],
                candidate["CANDIDATE NAME"],
                candidate["SOURCE"],
                candidate["APPLIED DATE"],
                candidate["RECRUITMENT PHASE\n(Pipeline)"],
                candidate["FINAL DECISION"],
                candidate["HR VIEW"],
                candidate["HIRING MANAGER VIEW"],
                candidate["DECISION MAKER VIEW"],
                candidate["COMMENTS"],
                candidate["RECIEVED APPLICATION COMMENTS"]
            ])
        
        print(f"✅ Added {len(candidates_df)} candidates")
    
    # Save
    wb.save(file_path)
    print(f"✅ Mock data saved to {file_path}")
    print(f"\n📊 Summary:")
    print(f"   - {len(jobs_df)} Job Openings")
    print(f"   - {len(candidates_df)} Candidates")
    print(f"   - {len(RECRUITERS)} Recruiters")
    print(f"   - {len(DEPARTMENTS)} Departments")

if __name__ == "__main__":
    file_path = "Luigi Recruitment-Tracker-Someka-Excel-Template-V9-Free-Version-2.xlsx"
    add_mock_data_to_excel(file_path, num_jobs=15, num_candidates=30)
